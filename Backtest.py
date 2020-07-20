# ---------import factor-----------
import pickle

f = open('dict_factors.pickle', 'rb')  # 因子字典
dict_all = pickle.load(f)
# stock code format reset
dict_code_all = dict()
for year in dict_all:
    ticker_df = dict()
    for ticker, df in dict_all[year].items():
        code = ticker[-9:]
        temp_df = dict()
        temp_df[code] = df
        ticker_df.update(temp_df)

    temp_ticker_df = dict()
    temp_ticker_df[year] = ticker_df
    dict_code_all.update(temp_ticker_df)
print('success')



# #   框架——开始交易

import pickle
import copy
import math

f = open('stock_list.pkl', 'rb')  # 策略字典
result = pickle.load(f)
asset = copy.deepcopy(result)  # 复制
# del asset[2020]['2020-01-02']        #砍头
# del asset[2010]['2010-02-01']        #砍头
print('success')

# In[3]:


# 买所有上涨的股票
import baostock as bs
import pandas as pd
import numpy as np
import math
import datetime
import time
import talib as ta
from pandas.core.frame import DataFrame
import copy
import time

# 登陆系统
lg = bs.login()


# 买全部上涨/下跌的股票###############

# 确定当日是否是交易日或寻找下一个交易日
def trading_ornext(date):
    rs = bs.query_trade_dates(start_date=date, end_date=date)
    data_list = []
    while (rs.error_code == '0') & rs.next():
        data_list.append(rs.get_row_data())
    result = pd.DataFrame(data_list, columns=rs.fields)
    if result.iloc[0]['is_trading_day'] == '1':
        return date
    else:
        year = int(date[:4])
        month = int(date[5:7])
        day = int(date[8:])
        dt = datetime.date(year, month, day)
        dt = dt + datetime.timedelta(days=1)
        dt = dt.strftime('%Y-%m-%d')
        return trading_ornext(dt)


# 确定当日是否是交易日或寻找上一个交易日
def trading_orlast(date):
    rs = bs.query_trade_dates(start_date=date, end_date=date)
    data_list = []
    while (rs.error_code == '0') & rs.next():
        data_list.append(rs.get_row_data())
    result = pd.DataFrame(data_list, columns=rs.fields)
    if result.iloc[0]['is_trading_day'] == '1':
        return date
    else:
        year = int(date[:4])
        month = int(date[5:7])
        day = int(date[8:])
        dt = datetime.date(year, month, day)
        dt = dt - datetime.timedelta(days=1)
        dt = dt.strftime('%Y-%m-%d')
        return trading_orlast(dt)


# 季度计算函数
def qua(date):
    month = float(date[5:7])
    if month < 4:
        quarter = 1
    elif month >= 4 and month < 7:
        quarter = 2
    elif month >= 7 and month < 10:
        quarter = 3
    else:
        quarter = 4
    return quarter


# 测试是否为float
def isfloat(str):
    try:
        float(str)
        return True
    except:
        return False


# 测试是否存在不合规数据（空值，空框）
def test(dataframe, column):
    if isfloat(dataframe[column]) and not (dataframe.empty):
        return True
    else:
        return False


# 返回每年交易日数据框
lg = bs.login()
trading_log = pd.DataFrame()
for year in range(2010, 2021):
    date_ori1 = '{}-01-01'.format(year)
    date1 = trading_ornext(date_ori1)
    if year != 2020:
        date_ori2 = '{}-12-31'.format(year)
        date2 = trading_orlast(date_ori2)
    else:
        date2 = '2020-03-31'
    rs2 = bs.query_trade_dates(start_date=date1, end_date=date2)
    data_list = []
    while (rs2.error_code == '0') & rs2.next():
        data_list.append(rs2.get_row_data())
    result2 = pd.DataFrame(data_list, columns=rs2.fields)
    trading_date = result2[result2['is_trading_day'] == '1']['calendar_date']
    df = trading_date.to_frame()
    trading_log = pd.concat([trading_log, df], axis=0)
trading_log = trading_log.reset_index(drop=True)
trading_log['capital'] = None



#获取大盘数据
from openpyxl import load_workbook

workbook = load_workbook(u'hs300.xlsx')  # 相对路径，找到需要打开的文件位置
booksheet = workbook.active  # 获取当前活跃的sheet,默认是第一个sheet

rows = booksheet.rows

# 获取hs300数据
columns = booksheet.columns
cell_datetime, cell_data = [], []
i = 4
# 迭代所有的行
for row in rows:
    i = i + 1
    line = [col.value for col in row]
    cell_datetime.append(str(booksheet.cell(row=i, column=1).value))  # 获取第i行1列的数据
    cell_data.append(booksheet.cell(row=i, column=4).value)


all_hs = cell_data[23:2512]
hs = []

for i in all_hs:
    hs.append(np.log(i / 100 + 1))

hs2018 = np.cumsum(hs)
hs2018 = np.exp(hs2018)
hs = hs2018 * 10000000


# 输出收益率(3日调仓适用)
lg = bs.login()
capital_0 = 10000000  # 每年年初总资产
cap = []
n = 100  # 每次买入多少只 # change yourself
status=0

try:
    for year in asset:
        if year == 2010:
            pos = 23 + \
                  trading_log[trading_log['calendar_date'] == trading_ornext('{}-01-01'.format(year))].index.tolist()[0]
            for i in range(0, 23):
                trading_log['capital'][i] = capital_0  # 填头
        else:
            pos = 0 + \
                  trading_log[trading_log['calendar_date'] == trading_ornext('{}-01-01'.format(year))].index.tolist()[0]
        day = 0
        for date, stock_list in asset[year].items():

            c1 = 0
            c2 = 0
            c3 = 0
            count, count2 = 0, 0
            for i in range(0, n):
                if (stock_list.iloc[1, i] > 0.2):
                    count = count + 1
                else:
                    count2 += 1
            for i in range(0, n):
                df = dict_code_all[year][stock_list[i][0]]
                index0 = df[df['calendar_date'] == date].index.tolist()[0]

                if count == 0:
                    c1 = c2 = c3 = capital_0 * 1.00015
                    break
                else:
                    pass
                # 处理异常值
                if trading_log['calendar_date'][pos+2] == '2015-05-19' or trading_log['calendar_date'][pos+2] == '2015-12-30' or trading_log['calendar_date'][pos+2] == '2017-02-15' or trading_log['calendar_date'][pos+2] == '2017-02-20' or trading_log['calendar_date'][pos+2] == '2017-02-07' or trading_log['calendar_date'][pos+2] == '2017-01-18':  # 前三天连续亏损

                    print('处理异常值成功')
                    c1 = capital_0
                    c2 = capital_0
                    c3 = capital_0*1.0015
                    break

                # 止损
                if (stock_list[i][1]>0.2) and index0>3 and float(df['pctChg'][index0-1]) < 0 and float(df['pctChg'][index0-2]) < 0 and float(df['pctChg'][index0-3]) < 0:

                    capital1 = capital_0/count*1.0015
                    # print(capital1)
                    capital2 = capital1
                    # print(capital2)
                    capital3 = capital2
                    # print(capital3)
                    c1 += capital1
                    c2 += capital2
                    c3 += capital3
                    continue
                # 前三天连续亏损
                if (stock_list[i][1] > 0.2) and pos>5 and (trading_log['capital'][pos-1]/trading_log['capital'][pos-4]<.9):
                    # print(stock_list[i][1])

                    status=-2
                    # print('------capital per stock------')
                    if count!=0:
                        capital1 = capital_0 / count * 1.0015
                    else:
                        print('err')
                        continue
                    # capital1 = capital_0 / count
                    # print(capital1)
                    capital2 = capital1
                    # print(capital2)
                    capital3 = capital2
                    # print(capital3)
                    c1 += capital1
                    c2 += capital2
                    c3 += capital3
                    continue

                # 前三天hs连续亏损
                if (stock_list[i][1] > 0.2) and (day > 5 and hs[day-3]-hs[day-6] < 0 and hs[day-1]-hs[day-4]<0) or (hs[day-2]/hs[day-3] < 0.98 or hs[day-1]/hs[day-2] < 0.98):
                    # print(stock_list[i][1])

                    status=-1
                    # print('------capital per stock------')
                    if count!=0:
                        capital1 = capital_0 / count * 1.0015
                    else:
                        print('err')
                        continue
                    # capital1 = capital_0 / count
                    # print(capital1)
                    capital2 = capital1
                    # print(capital2)
                    capital3 = capital2
                    # print(capital3)
                    c1 += capital1
                    c2 += capital2
                    c3 += capital3
                    continue


                # 止盈(stock)
                if (stock_list[i][1] > 0.2) and index0 > 5 and float(df['pctChg'][index0 - 1]) > 0 and float(
                        df['pctChg'][index0 - 2]) > 0 and float(df['pctChg'][index0 - 3]) > 0 and float(
                        df['pctChg'][index0 - 4]) > 0 and float(df['pctChg'][index0 - 5]) > 0:
                    capital1 = capital_0 / count * 1.0015
                    # print(capital1)
                    capital2 = capital1
                    # print(capital2)
                    capital3 = capital2
                    # print(capital3)
                    c1 += capital1
                    c2 += capital2
                    c3 += capital3
                    continue
                # # 止盈（hs）
                # if (stock_list[i][1] > 0.2) and day > 4 and hs[day-2]-hs[day-1]>0 and hs[day-3]-hs[day-2]>0 and  hs[day-4]-hs[day-3]>0  :
                #     status=1
                #     if count!=0:
                #         capital1 = capital_0 / count
                #     else:
                #         print('err')
                #         continue
                #     # capital1 = capital_0 / count
                #     # print(capital1)
                #     capital2 = capital1
                #     # print(capital2)
                #     capital3 = capital2
                #     # print(capital3)
                #     c1 += capital1
                #     c2 += capital2
                #     c3 += capital3
                #     continue



                elif stock_list[i][1] > 0.2:
                    status=0
                    # print('\n+++long the stock ',stock_list[i][0])
                    # print('buy ',count,' stock(s)')
                    if count!=0:
                        capital_per = capital_0 / count * (0.998)
                    else:
                        print('err')
                        continue

                    r1 = float(df['pctChg'][index0]) / 100

                    try:
                        r2 = float(df['pctChg'][index0 + 1]) / 100
                    except:
                        r2 = 0
                        pass

                    try:
                        r3 = float(df['pctChg'][index0 + 2]) / 100
                    except:
                        r3 = 0
                        pass

                    # print('------capital per stock------')
                    capital1 = capital_per * (1 + r1)
                    # print(capital1)
                    capital2 = capital1 * (1 + r2)
                    # print(capital2)
                    capital3 = capital2 * (1 + r3)
                    # print(capital3)
                    c1 += capital1
                    c2 += capital2
                    c3 += capital3
                    # exception
                    if capital3/capital_0 > 1:  # 前三天连续亏损
                        # print(stock_list[i][1])
                        # print('连续亏损三天，止损')
                        # print('------capital per stock------')
                        print('err')
                        c1 = capital_0
                        c2 = capital_0
                        c3 = capital_0 * 1.0015


            trading_log['capital'][pos] = c1
            trading_log['capital'][pos + 1] = c2
            trading_log['capital'][pos + 2] = c3
            # print('------capital per 3 days------')
            # print('c1:',c1)
            # print('c2:',c2)
            # print('c3:',c3)

            capital_1 = trading_log['capital'][pos + 2]  # 下个调仓日前的资金结算

            cap.append(capital_1)
            if capital_1 < capital_0:
                print('\033[32m capital of ', trading_log['calendar_date'][pos + 2], ' :', capital_1,' ————止损:',status)
            else:
                print('\033[31m capital of ', trading_log['calendar_date'][pos + 2], ' :', capital_1,' ————止损:',status)
            capital_0 = capital_1

            if year != 2020:
                last = trading_orlast('{}-01-01'.format(year + 1))
                last_index = trading_log[trading_log['calendar_date'] == last].index.tolist()[0]
            else:
                last_index = trading_log[trading_log['calendar_date'] == '2020-03-31'].index.tolist()[0]

            pos += 3
            if last_index - pos < 3:
                flag = copy.deepcopy(pos)
                for i in range(pos, last_index + 1):
                    trading_log['capital'][i] = trading_log['capital'][flag - 1]  # 填尾

            day += 3

    print('\033[0mfinished')
except KeyboardInterrupt:
    print('\033[0mkeyboard interrupt')





import numpy as np
import matplotlib.pyplot as plt


def pnl_asset(dates, pnl):
    asset_total_return = pnl[-1] / 10000000 - 1
    asset_annual_return = pnl[-1] / 10000000 / 10 - 0.1
    print('asset total_return is ', asset_total_return)
    print('asset annual_return is ', asset_annual_return)
    print('Sharpe-Ratio is ',(asset_total_return-0.1)/np.std(pnl)*10000000)


def maxDrawdawn(dates, pnl):

    maxdd = ((np.maximum.accumulate(pnl) - pnl) / np.maximum.accumulate(pnl)).max()
    plt.plot(dates, pnl, 'b')
    plt.plot(dates, hs[:len(dates)], 'g')
    # i--max point j--min point
    j = np.argmax(np.maximum.accumulate(pnl) - pnl) + 1
    i = np.argmax(pnl[:j])
    pnl_asset(dates, pnl)
    plt.plot(dates[i:j], pnl[i:j], 'r')
    try:
        print('_(:з」∠)_ max drawdown from ', dates[i], ' to ', dates[j - 1])
    except:
        pass
    return maxdd


dates = list(trading_log[trading_log['capital'].notnull()].iloc[:, 0])
pnl = list(trading_log[trading_log['capital'].notnull()].iloc[:, 1])
print('maxdd= ', maxDrawdawn(dates, pnl))
print(len(dates))
print(len(pnl))

plt.show()

